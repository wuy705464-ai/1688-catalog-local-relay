"""Export AI-selected local products into customer-facing catalog workbooks.

The exporter is strict by default: a row is written only when exactly three
selected image records have the same offer_id and record-hash directory as the
product. Any mismatch aborts the export instead of silently pairing wrong data.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import uuid
from copy import copy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.views import Selection
from PIL import Image, ImageOps


ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.local_store import LocalStore  # noqa: E402
from src.runtime_config import load_runtime_config, resolve_project_path  # noqa: E402


def safe_name(value: str) -> str:
    cleaned = re.sub(r"[^\w\-\u4e00-\u9fff]+", "_", value or "all").strip("_")
    return cleaned[:60] or "all"


def chunks(values: List[Dict[str, Any]], size: int) -> Iterable[List[Dict[str, Any]]]:
    for index in range(0, len(values), size):
        yield values[index:index + size]


def validate_alignment(product: Dict[str, Any], selected_count: int = 3) -> List[Path]:
    offer_id = str(product["offer_id"])
    hash_prefix = str(product["record_hash"])[:16]
    images = product.get("selected_images") or []
    if len(images) != selected_count:
        raise ValueError(f"offer {offer_id}: expected {selected_count} selected images, got {len(images)}")
    ranks = [int(image["selected_rank"]) for image in images]
    if ranks != list(range(1, selected_count + 1)):
        raise ValueError(f"offer {offer_id}: selected ranks are not 1..{selected_count}")

    paths: List[Path] = []
    for image in images:
        if str(image["offer_id"]) != offer_id:
            raise ValueError(f"offer {offer_id}: image foreign key mismatch")
        path = Path(str(image["local_path"])).resolve()
        if offer_id not in path.parts or hash_prefix not in path.parts:
            raise ValueError(f"offer {offer_id}: image path is outside its versioned product folder")
        if not path.exists():
            raise FileNotFoundError(f"offer {offer_id}: selected image missing: {path}")
        paths.append(path)
    return paths


def _fit_tile(path: Path, size: Tuple[int, int]) -> Image.Image:
    with Image.open(path) as opened:
        image = ImageOps.exif_transpose(opened).convert("RGB")
    image.thumbnail(size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGB", size, "white")
    canvas.paste(image, ((size[0] - image.width) // 2, (size[1] - image.height) // 2))
    return canvas


def make_triptych(image_paths: List[Path], output: Path, size: Tuple[int, int] = (900, 450)) -> Path:
    if len(image_paths) != 3:
        raise ValueError("triptych requires exactly three images")
    tile_w = size[0] // 3
    canvas = Image.new("RGB", size, "white")
    for index, path in enumerate(image_paths):
        tile = _fit_tile(path, (tile_w, size[1]))
        canvas.paste(tile, (index * tile_w, 0))
        if index:
            for y in range(size[1]):
                canvas.putpixel((index * tile_w, y), (225, 225, 225))
    output.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output, "JPEG", quality=86, optimize=True)
    return output


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        target.alignment = copy(source.alignment)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(_display_value(item) for item in value if _display_value(item))
    if isinstance(value, dict):
        return ", ".join(f"{key}: {_display_value(item)}" for key, item in value.items())
    return str(value).strip()


def format_specs(record: Dict[str, Any], max_items: int = 8) -> Tuple[str, str]:
    specs = record.get("specs") if isinstance(record.get("specs"), dict) else {}
    lines: List[str] = []
    material = ""
    for key, value in specs.items():
        label = str(key).strip()
        shown = _display_value(value)
        if not label or not shown:
            continue
        if not material and ("材质" in label or "material" in label.lower()):
            material = shown
        if len(lines) < max_items:
            lines.append(f"• {label}: {shown}")
    return "\n".join(lines), material


def export_one(
    template: Path,
    products: List[Dict[str, Any]],
    output_path: Path,
    cfg: Dict[str, Any],
    freeze_panes: bool = True,
) -> Tuple[Path, Path]:
    if not products:
        raise ValueError("no products to export")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb["Product Catalog"]
    export_cfg = cfg.get("catalog_export", {})
    display_size = tuple(export_cfg.get("embedded_image_size_px", [300, 150]))
    row_height = float(export_cfg.get("row_height_points", 125))
    # Row 3 in the template is an example row. Preserve its style, but remove
    # all template data/drawings so customer files contain only real products.
    start_row = 3
    manifest: List[Dict[str, Any]] = []

    ws._images = []
    for existing_row in range(start_row, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            cell = ws.cell(existing_row, column)
            cell.value = None
            cell.hyperlink = None
            cell.comment = None

    temp_dir = output_path.parent / f".catalog-triptych-{uuid.uuid4().hex}"
    temp_dir.mkdir(parents=True, exist_ok=False)
    try:
        for index, product in enumerate(products):
            row = start_row + index
            if row > ws.max_row:
                _copy_row_style(ws, start_row, row)
            image_paths = validate_alignment(product, selected_count=3)
            collage = make_triptych(image_paths, temp_dir / f"{product['offer_id']}.jpg")

            record = product["record"]
            specs_text, material = format_specs(record)
            ws.row_dimensions[row].height = row_height
            ws.cell(row=row, column=1, value=index + 1)
            ws.cell(row=row, column=2, value=product.get("category") or "")
            ws.cell(row=row, column=3, value=f"1688-{product['offer_id']}")
            ws.cell(row=row, column=4, value=product.get("title") or "")
            ws.cell(row=row, column=6, value=specs_text)
            ws.cell(row=row, column=7, value=product.get("price_display") or "")
            ws.cell(row=row, column=11, value=material)
            ws.cell(row=row, column=12, value=product.get("size_raw") or "")

            picture = XLImage(str(collage))
            picture.width, picture.height = display_size
            ws.add_image(picture, f"E{row}")
            manifest.append(
                {
                    "row": row,
                    "offer_id": product["offer_id"],
                    "record_hash": product["record_hash"],
                    "title": product.get("title") or "",
                    "category": product.get("category") or "",
                    "price": product.get("price_display") or "",
                    "size": product.get("size_raw") or "",
                    "selected_images": [
                        {
                            "rank": image["selected_rank"],
                            "sha256": image["sha256"],
                            "role": image["role"],
                            "source_url": image["source_url"],
                        }
                        for image in product["selected_images"]
                    ],
                    "source_url": record.get("url", ""),
                }
            )
        last_row = start_row + len(products) - 1
        if freeze_panes:
            ws.freeze_panes = "A3"
            # The template freezes both columns and rows. openpyxl keeps the old
            # top-right/bottom-right selections when changing it to a row-only
            # freeze, which makes desktop Excel repair sheet1.xml on open.
            ws.sheet_view.selection = [
                Selection(pane="bottomLeft", activeCell="A3", sqref="A3")
            ]
        else:
            # Collection mode stays fully scrollable/editable. Re-export with
            # freeze_panes=True only after the crawl is complete.
            for sheet in wb.worksheets:
                sheet.freeze_panes = None
                sheet.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
            ws.sheet_view.selection = [Selection(activeCell="A3", sqref="A3")]
        ws.print_area = f"A1:L{last_row}"
        ws.sheet_view.showGridLines = False
        wb.save(output_path)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    manifest_path = output_path.with_suffix(".manifest.json")
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path, manifest_path


def export_master_atomic(
    template: Path,
    products: List[Dict[str, Any]],
    output_path: Path,
    cfg: Dict[str, Any],
    finalize: bool = False,
) -> Tuple[Path, Path]:
    """Replace one master workbook only after a complete aligned export.

    The staging file keeps the existing master intact if validation or image
    embedding fails. On Windows, Excel must be closed before the final replace;
    frozen panes do not control the operating-system file lock.
    """
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    staging_path = output_path.with_name(f".{output_path.stem}.{uuid.uuid4().hex}.xlsx")
    staging_manifest = staging_path.with_suffix(".manifest.json")
    final_manifest = output_path.with_suffix(".manifest.json")
    try:
        export_one(
            template=template,
            products=products,
            output_path=staging_path,
            cfg=cfg,
            freeze_panes=finalize,
        )
        staging_path.replace(output_path)
        staging_manifest.replace(final_manifest)
    except PermissionError as exc:
        raise PermissionError(
            f"Close the master workbook in Excel before syncing: {output_path}"
        ) from exc
    finally:
        staging_path.unlink(missing_ok=True)
        staging_manifest.unlink(missing_ok=True)
    return output_path, final_manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Export customer catalogs from the local SQLite database")
    parser.add_argument("--config", type=Path, default=None)
    parser.add_argument("--db", type=Path, default=None)
    parser.add_argument("--template", type=Path, default=None)
    parser.add_argument("--out-dir", type=Path, default=ROOT / "output" / "customer_catalogs")
    parser.add_argument("--category", default="")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--all-categories-together", action="store_true")
    parser.add_argument(
        "--master-output",
        type=Path,
        default=None,
        help="Atomically sync all ready products into this one master workbook",
    )
    parser.add_argument(
        "--finalize",
        action="store_true",
        help="Freeze the first two rows in master output after crawling is complete",
    )
    args = parser.parse_args()

    if args.finalize and args.master_output is None:
        parser.error("--finalize requires --master-output")

    cfg = load_runtime_config(args.config)
    relay_cfg = cfg.get("relay", {})
    db_path = args.db.resolve() if args.db else resolve_project_path(str(relay_cfg.get("database") or ""), "data/catalog.db")
    template = args.template.resolve() if args.template else resolve_project_path(
        str(cfg.get("output", {}).get("template") or "template.xlsx"), "template.xlsx"
    )
    store = LocalStore(db_path)
    products = store.products_for_export(category=args.category, limit=args.limit)
    if not products:
        raise SystemExit("No AI-selected products are ready for export")

    if args.master_output is not None:
        workbook_path, manifest_path = export_master_atomic(
            template=template,
            products=products,
            output_path=args.master_output,
            cfg=cfg,
            finalize=args.finalize,
        )
        print(f"catalog:  {workbook_path}")
        print(f"manifest: {manifest_path}")
        return

    grouped: Dict[str, List[Dict[str, Any]]] = {}
    if args.all_categories_together:
        grouped["all"] = products
    else:
        for product in products:
            grouped.setdefault(product.get("category") or "uncategorized", []).append(product)

    max_per_file = int(cfg.get("catalog_export", {}).get("max_products_per_file", 150))
    outputs = []
    for category, rows in grouped.items():
        for part, subset in enumerate(chunks(rows, max_per_file), start=1):
            output_path = args.out_dir.resolve() / f"customer_catalog_{safe_name(category)}_{part:03d}.xlsx"
            workbook_path, manifest_path = export_one(template, subset, output_path, cfg)
            outputs.append((workbook_path, manifest_path))
    for workbook_path, manifest_path in outputs:
        print(f"catalog:  {workbook_path}")
        print(f"manifest: {manifest_path}")


if __name__ == "__main__":
    main()
