from __future__ import annotations

import json
import shutil
import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from scripts.export_customer_catalog import export_master_atomic, export_one  # noqa: E402
from src.local_store import LocalStore  # noqa: E402


def record(offer_id: str, title: str, category: str):
    return {
        "schema_version": 3,
        "offer_id": offer_id,
        "url": f"https://detail.1688.com/offer/{offer_id}.html",
        "title": title,
        "category": category,
        "price": {"raw": "¥3.00", "display": "¥3.00", "tiers": []},
        "size": {"raw": "45cm", "source": "attribute"},
        "specs": {"链长": "45cm"},
        "image_urls": [f"https://img.alicdn.com/{offer_id}/{i}.jpg" for i in range(8)],
        "collected_at": "2026-08-21T12:00:00Z",
    }


class CatalogExportTests(unittest.TestCase):
    def test_two_products_keep_text_and_images_on_the_same_rows(self):
        temp_root = ROOT / "tests" / "_runtime"
        temp_root.mkdir(parents=True, exist_ok=True)
        temp = temp_root / f"catalog-export-{uuid.uuid4().hex}"
        temp.mkdir()
        try:
            store = LocalStore(temp / "catalog.db")
            expected = [
                ("940000000001", "红色项链", "项链", (220, 30, 30)),
                ("940000000002", "蓝色手链", "手链", (30, 30, 220)),
            ]
            for offer_id, title, category, color in expected:
                store.upsert_product(record(offer_id, title, category))
                job = store.claim_next_job("export-test")
                version_dir = temp / "products" / "selected" / offer_id / job["version_hash"][:16]
                version_dir.mkdir(parents=True)
                candidates = []
                for rank in range(1, 4):
                    path = version_dir / f"{rank:02d}.jpg"
                    Image.new("RGB", (400, 400), color).save(path)
                    candidates.append(
                        {
                            "candidate_index": rank,
                            "source_url": f"https://img.alicdn.com/{offer_id}/{rank}.jpg",
                            "width": 400,
                            "height": 400,
                            "sha256": f"{offer_id}-{rank}",
                            "role": ("hero", "detail", "lifestyle")[rank - 1],
                            "reason": "test",
                            "is_selected": True,
                            "selected_rank": rank,
                            "local_path": str(path),
                        }
                    )
                self.assertTrue(store.complete_selection(offer_id, job["version_hash"], candidates, "test", 3))

            products = store.products_for_export()
            output = temp / "customer_catalog.xlsx"
            manifest = temp / "customer_catalog.manifest.json"
            cfg = {"catalog_export": {"embedded_image_size_px": [300, 150], "row_height_points": 125}}
            workbook_path, manifest_path = export_one(ROOT / "template.xlsx", products, output, cfg)
            self.assertEqual(manifest_path, manifest)

            wb = load_workbook(workbook_path)
            ws = wb["Product Catalog"]
            row_pairs = {ws["C3"].value: ws["D3"].value, ws["C4"].value: ws["D4"].value}
            self.assertEqual(
                row_pairs,
                {"1688-940000000001": "红色项链", "1688-940000000002": "蓝色手链"},
            )
            self.assertEqual(len(ws._images), 2)
            self.assertEqual([image.anchor._from.row for image in ws._images], [2, 3])
            self.assertIn("链长: 45cm", ws["F3"].value)
            self.assertEqual(ws.freeze_panes, "A3")
            self.assertEqual(len(ws.sheet_view.selection), 1)
            self.assertEqual(ws.sheet_view.selection[0].pane, "bottomLeft")
            self.assertEqual(ws.sheet_view.selection[0].sqref, "A3")

            rows = json.loads(manifest.read_text(encoding="utf-8"))
            self.assertEqual({row["offer_id"] for row in rows}, {"940000000001", "940000000002"})
            self.assertEqual({row["row"] for row in rows}, {3, 4})
            for row in rows:
                self.assertTrue(all(row["offer_id"] in image["sha256"] for image in row["selected_images"]))

            master_template = temp / "master.xlsx"
            shutil.copy2(ROOT / "template.xlsx", master_template)
            master_path, master_manifest = export_master_atomic(
                master_template,
                products,
                master_template,
                cfg,
                finalize=False,
            )
            self.assertEqual(master_path, master_template.resolve())
            self.assertTrue(master_manifest.exists())
            master_wb = load_workbook(master_path)
            self.assertIsNone(master_wb["Product Catalog"].freeze_panes)
            self.assertIsNone(master_wb["How to Use"].freeze_panes)
            self.assertEqual(
                {master_wb["Product Catalog"]["C3"].value, master_wb["Product Catalog"]["C4"].value},
                {"1688-940000000001", "1688-940000000002"},
            )

            export_master_atomic(
                ROOT / "template.xlsx",
                products,
                master_template,
                cfg,
                finalize=True,
            )
            finalized_wb = load_workbook(master_template)
            self.assertEqual(finalized_wb["Product Catalog"].freeze_panes, "A3")
        finally:
            shutil.rmtree(temp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
